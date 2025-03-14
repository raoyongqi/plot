import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Create the data from the LaTeX table
data = {
    'spec.variable': [
        'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 
        'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load',
        'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load', 'Pathogen Load'],
    'env.variable': [
        'SRAD', 'S SAND', 'LON', 'PSEA', 'WIND', 'VAPR', 'MAX MAT', 'ELEV', 'TSEA', 'MIN MAT', 'LAT', 'MAP',
        'AVG MAT', 'T SAND', 'T BULK DEN', 'T REF BULK', 'S CLAY'],
    'Mantel r': [
        0.06, 0.08, -0.00, -0.00, 0.00, -0.03, -0.00, -0.02, -0.02, -0.02, -0.01, -0.02, -0.01, 0.03, -0.01, 
        0.05, 0.08],
    'p-value': [
        0.0, 0.0, 0.04, 0.80, 0.38, 0.0, 0.25, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
}

df = pd.DataFrame(data)

file_path = "mantel_results.xlsx"
df.to_excel(file_path, index=False)

wb = load_workbook(file_path)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
    for cell in row:
        if cell.value < 0.05:
            cell.font = Font(bold=True)

wb.save(file_path)

print(f"File saved with bolded p-values at: {file_path}")
